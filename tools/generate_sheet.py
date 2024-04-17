import json
import sys
import os
import subprocess
import time
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

sys.path.append("/".join(__file__.split("/")[0:-2]))
# pylint: disable-next=wrong-import-position
from lib.aux import orange, purple


components_keys = [
    "abastecimento-de-agua",
    "esgotamento-sanitario",
    "manejo-das-aguas-pluviais",
    "manejo-de-residuos-solidos",
]
components_uppercase = [
    "ABASTECIMENTO DE ÁGUA",
    "ESGOTAMENTO SANITÁRIO",
    "MANEJO DAS ÁGUAS PLUVIAIS",
    "MANEJO DE RESÍDUOS SÓLIDOS",
]
components_acronym = [
    "AA",
    "ES",
    "AP",
    "RS",
]


def merge_and_center(tab, start_row, end_row, start_column, end_column=None):
    tab.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column if end_column else start_column,
    )
    tab.cell(row=start_row, column=start_column).alignment = Alignment(
        vertical="center", horizontal="center"
    )


def paint_cells(tab, color, row, column):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell = tab.cell(row=row, column=column)
    cell.fill = fill


def set_border(tab, row, column):
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    cell = tab.cell(row=row, column=column)
    cell.border = border


def set_header(tab, aux, key):

    component_uppercase = components_uppercase[components_keys.index(key)]
    row = aux + 3

    if tab.title == "Quadro3.3":
        tab.cell(row=row + 1, column=1).value = "OBJETIVO"
        tab.cell(row=row + 1, column=2).value = "PROJETO"
        tab.cell(row=row + 1, column=3).value = "DESCRIÇÃO DA AÇÃO PROPOSTA"
        tab.cell(row=row + 1, column=4).value = "HORIZONTE DO PMSB (anos)"

    elif tab.title == "Quadro3.4":
        tab.cell(row=row + 1, column=3).value = "META / PRAZO"
        tab.cell(row=row + 1, column=4).value = "ANDAMENTO DA AÇÃO"
        tab.cell(row=row + 1, column=24).value = "TIPO DE PROBLEMA"
        tab.cell(row=row + 1, column=25).value = "MOTIVO"
        tab.cell(row=row + 1, column=26).value = "JUSTIFICATIVA"

        for i in range(3):
            merge_and_center(tab, row + 1, row + 2, 24 + i)

    elif tab.title == "Quadro3.5":
        tab.cell(row=row + 1, column=3).value = "TIPO DE PROBLEMA"
        tab.cell(row=row + 1, column=4).value = "MOTIVO PROPOSTA DE ALTERNATIVA DE AÇÃO"
        tab.cell(row=row + 1, column=5).value = "RESPONSÁVEL"
        tab.cell(row=row + 1, column=6).value = "PARCERIAS"
        tab.cell(row=row + 1, column=7).value = "IMPACTOS"
        tab.cell(row=row + 2, column=7).value = "PRAZO"
        tab.cell(row=row + 2, column=8).value = "CUSTO"

        merge_and_center(tab, row + 1, row + 1, 7, 8)
        for i in range(3):
            merge_and_center(tab, row + 1, row + 2, i + 3)

    columns = {"Quadro3.3": 23, "Quadro3.4": 26, "Quadro3.5": 8}
    merge_and_center(tab, row, row, 1, columns[tab.title])
    for i in range(columns[tab.title]):
        set_border(tab, row, i + 1)
        set_border(tab, row + 1, i + 1)
        set_border(tab, row + 2, i + 1)
        paint_cells(tab, "B4C6E7", row + 1, i + 1)
        paint_cells(tab, "B4C6E7", row + 2, i + 1)

    # Common to Tab3.3 and Tab3.4
    if tab.title != "Quadro3.5":
        merge_and_center(tab, row + 1, row + 1, 4, 23)
        for i in range(20):
            tab.cell(row=row + 2, column=i + 4).value = i + 1
    tab.cell(row=row, column=1).value = "COMPONENTE: " + component_uppercase

    # Common to Tab3.4 and Tab3.5
    if tab.title != "Quadro3.3":
        tab.cell(row=row + 1, column=1).value = "DESCRIÇÃO DA AÇÃO"
        tab.cell(row=row + 1, column=2).value = "CLASSIFICAÇÃO DA AÇÃO"

    # Common to all tabs
    merge_and_center(tab, row + 1, row + 2, 1)
    merge_and_center(tab, row + 1, row + 2, 2)
    merge_and_center(tab, row + 1, row + 2, 3)
    paint_cells(tab, "8EAADB", row, 1)


def convert_investment(investment):
    if not investment:
        return None
    if not "mil" in investment:
        try:
            return float(
                investment.replace("R$", "")
                .replace(" ", "")
                .replace(".", "")
                .replace(",", ".")
            )
        except ValueError:
            return None
    else:
        pattern = r"(?:(?:R\$\s*)?(\d[\d\.,]*)\s*(mil(?:hões?|ha)?)|(\d+(?:\.\d+)?))"
        match = re.match(pattern, investment, re.IGNORECASE)

        if match:
            number_str = match.group(1) or match.group(3)

            if number_str:
                number_str = number_str.replace(",", ".")

                if "milh" in investment.lower():
                    return float(number_str) * 1_000_000
                elif "mil" in investment.lower():
                    return float(number_str) * 1_000
                else:
                    return float(number_str)

        return None


def get_actions_index(deadline):

    if "Imediato" in deadline or "Emergencial" in deadline:
        return 1
    elif "Curto" in deadline:
        return 2
    elif "Médio" in deadline:
        return 3
    elif "Longo" in deadline:
        return 4
    else:
        print(
            f"Prazo {orange(deadline)} não identificado, assumindo como {purple('Imediato')}"
        )
        return 1


def generate_tab_3_1(data_json, tab):

    aux = 0

    for key in components_keys:
        component_uppercase = components_uppercase[components_keys.index(key)]
        for i, value in enumerate(data_json[key]):

            deadline = value.get("deadline")
            objective = value.get("objective")
            investment = convert_investment(value.get("investment"))
            count_objectives = len(data_json[key])
            start_row = aux + 2
            end_row = aux + 1 + len(data_json[key])
            row = i + start_row

            tab.cell(row=row, column=1).value = component_uppercase
            tab.cell(row=row, column=2).value = count_objectives
            tab.cell(row=row, column=3).value = objective

            if deadline:
                tab.cell(row=row, column=4).value = deadline
            if investment:
                tab.cell(row=row, column=5).value = investment

        merge_and_center(tab, start_row, end_row, 1)
        merge_and_center(tab, start_row, end_row, 2)
        aux += len(data_json[key])


def generate_tab_3_2(data_json, tab):

    header_rows = 4  # rows on header plus 1

    for key in components_keys:

        row = components_keys.index(key) + header_rows
        count_actions_splited = [0, 0, 0, 0, 0]
        count_investiment_splited = [0, 0, 0, 0, 0]

        for _, value in enumerate(data_json[key]):

            actions = value.get("actions")
            deadline = value.get("deadline")
            investment = convert_investment(value.get("investment"))
            if not actions or not deadline:
                continue

            index = get_actions_index(deadline)
            count_actions_splited[0] += len(actions)
            count_actions_splited[index] += len(actions)
            count_investiment_splited[index] += investment if investment else 0

        tab.cell(row=row, column=2).value = count_actions_splited[0]
        for i in range(4):
            tab.cell(row=row, column=(i * 2) + 3).value = count_actions_splited[i + 1]
            tab.cell(row=row, column=(i * 2) + 4).value = count_investiment_splited[
                i + 1
            ]


def generate_tab_3_3(data_json, tab):

    aux = -2
    header_rows = 4  # rows on header plus 1

    for key in components_keys:

        set_header(tab, aux, key)
        aux += header_rows

        for i, value in enumerate(data_json[key]):

            actions = value.get("actions")

            if not actions:
                continue

            start_row = aux + i + 2
            end_row = aux + i + 1 + len(actions)

            for j, action in enumerate(actions):

                row = j + start_row
                objective = value.get("objective")

                tab.cell(row=row, column=1).value = objective
                tab.cell(row=row, column=3).value = action

                set_border(tab, row, 1)
                set_border(tab, row, 2)
                set_border(tab, row, 3)

                for k in range(20):
                    set_border(tab, row, header_rows + k)

            merge_and_center(tab, start_row, end_row, 1)
            aux += len(actions) - 1

        aux += len(data_json[key])


def generate_tab_3_4_and_3_5(data_json, tab):

    aux = -2
    aux2 = 0
    header_rows = 4  # rows on header plus 1

    for key in components_keys:

        set_header(tab, aux, key)
        aux += header_rows

        for i, value in enumerate(data_json[key]):

            actions = value.get("actions")

            if not actions:
                continue

            for j, data in enumerate(actions):

                row = i + j + aux + 2
                action_global_index = f"A{j + aux2 + 1}"
                cmp_acronym = components_acronym[components_keys.index(key)]
                objective_index = f"O{i + 1}"
                action_index = f"a{j + 1}-{len(actions)}"

                tab.cell(row=row, column=1).value = data
                # TODO: Implement "Projeto" at "CLASSIFICAÇÃO DA AÇÃO", ex: /P1/
                tab.cell(row=row, column=2).value = (
                    f"{action_global_index}/{cmp_acronym}/{objective_index}/{action_index}"
                )
                if tab.title == "Quadro3.4":
                    tab.cell(row=row, column=3).value = "Programado"
                    tab.cell(row=row + 1, column=3).value = "Executado"

                    merge_and_center(tab, row, row + 1, 1)
                    merge_and_center(tab, row, row + 1, 2)
                    merge_and_center(tab, row, row + 1, 24)
                    merge_and_center(tab, row, row + 1, 25)
                    merge_and_center(tab, row, row + 1, 26)

                    for k in range(26):
                        set_border(tab, row, k + 1)
                        set_border(tab, row + 1, k + 1)
                else:
                    for k in range(8):
                        set_border(tab, row, k + 1)

                aux += 1 if tab.title == "Quadro3.4" else 0
            aux2 += len(actions)
            aux += len(actions) - 1
        aux += len(data_json[key])


if __name__ == "__main__":

    with open("data/funasa.json", encoding="utf-8") as file:
        funasa = json.load(file)
    with open("cfg.json", encoding="utf-8") as f:
        excel_file_path = json.load(f)["excel_file_path"]

    workbook = load_workbook(excel_file_path)

    generate_tab_3_1(funasa, workbook["Quadro3.1"])
    generate_tab_3_2(funasa, workbook["Quadro3.2"])
    generate_tab_3_3(funasa, workbook["Quadro3.3"])
    generate_tab_3_4_and_3_5(funasa, workbook["Quadro3.4"])
    generate_tab_3_4_and_3_5(funasa, workbook["Quadro3.5"])

    workbook.save("tmp/sheet.xlsx")

    try:
        # fecha o LibreOffice se já estiver aberto
        subprocess.Popen(["pkill", "soffice"])
        # sleep para dar tempo de fechar o LibreOffice
        time.sleep(1)
        with open(os.devnull, "w", encoding="utf-8") as devnull:
            # TODO: Usar excel no lugar do libreoffice
            subprocess.Popen(
                ["libreoffice", "tmp/sheet.xlsx"], stdout=devnull, stderr=devnull
            )
    except Exception as e:
        print("Erro ao abrir o LibreOffice:", e)
